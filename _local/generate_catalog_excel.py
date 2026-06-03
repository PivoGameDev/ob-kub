import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
SUBHEADER_FONT = Font(bold=True, size=11)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

COLUMNS = [
    ('Категория', 22),
    ('Тип оборудования', 36),
    ('Объём, л', 12),
    ('Диаметр, мм', 13),
    ('Высота, мм', 13),
    ('Толщина стенки, мм', 16),
    ('Вес, кг', 11),
    ('Мощность, кВт', 14),
    ('Полный объём, л', 15),
    ('Рабочий объём, л', 16),
    ('Рубашки', 10),
    ('Давление, бар', 14),
    ('Цена, ₽', 16),
    ('Цена для утверждения, ₽', 24),
]

# ── helpers ──────────────────────────────────────────────────────────
def _specs(ref, target_vol, ref_vol, price_mult=1.0):
    ratio = target_vol / ref_vol
    cr = ratio ** (1/3)
    sr = ratio ** (2/3)
    d = max(200, round(ref['diameter'] * cr))
    h = max(300, round(ref['height'] * cr))

    if target_vol < 500:
        wall = 2
    elif target_vol < 3000:
        wall = 3
    elif target_vol < 10000:
        wall = 3
    elif target_vol < 50000:
        wall = 4
    elif target_vol < 100000:
        wall = 5
    else:
        wall = 6
    wall = max(wall, ref.get('wall', 2))

    weight = max(round(ref['weight'] * sr), 10)
    power = round(ref['power'] * sr, 1)

    price_scale = ratio ** 0.78
    price = max(round(ref['price'] * price_scale / 500) * 500, 5000)
    price = round(price * price_mult)

    full_vol = round(math.pi * (d / 2) ** 2 * h / 1000)
    work_vol = round(full_vol * 0.8)

    return {
        'diameter': d, 'height': h, 'wall': wall,
        'weight': weight, 'power': power, 'price': price,
        'full_volume': full_vol, 'working_volume': work_vol,
    }


def write_sheet(ws, title, rows, columns=COLUMNS):
    ws.title = title
    # header
    for ci, (col_name, col_w) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = col_w

    # data
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if ci == 1:
                cell.font = Font(bold=True)
            if ci >= 13 and isinstance(val, (int, float)) and val:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            if ci == 14:
                # empty approval column — light yellow fill
                cell.fill = PatternFill('solid', fgColor='FFF2CC')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


# ═══════════════════════════════════════════════════════════════════
#  1. BEER – CCT
# ═══════════════════════════════════════════════════════════════════
cct_data = {
    100:   {'title':'ЦКТ 100 л','diameter':400,'height_cyl':600,'height_cone':290,'height_full':1040,'full_volume':88,'working_volume':70,'wall':2,'jackets':2,'weight':45,'pressure':2.5,'from_price':58000},
    250:   {'title':'ЦКТ 250 л','diameter':500,'height_cyl':900,'height_cone':360,'height_full':1410,'full_volume':200,'working_volume':160,'wall':2,'jackets':2,'weight':65,'pressure':2.5,'from_price':94000},
    500:   {'title':'ЦКТ 500 л','diameter':650,'height_cyl':1100,'height_cone':470,'height_full':1720,'full_volume':417,'working_volume':334,'wall':2,'jackets':2,'weight':95,'pressure':2.5,'from_price':150000},
    1000:  {'title':'ЦКТ 1000 л','diameter':800,'height_cyl':1400,'height_cone':570,'height_full':2170,'full_volume':799,'working_volume':639,'wall':3,'jackets':2,'weight':160,'pressure':2.5,'from_price':247000},
    1500:  {'title':'ЦКТ 1500 л','diameter':900,'height_cyl':1600,'height_cone':640,'height_full':2440,'full_volume':1154,'working_volume':923,'wall':3,'jackets':2,'weight':210,'pressure':2.5,'from_price':338000},
    2000:  {'title':'ЦКТ 2000 л','diameter':1000,'height_cyl':1800,'height_cone':710,'height_full':2710,'full_volume':1600,'working_volume':1280,'wall':3,'jackets':2,'weight':270,'pressure':2.5,'from_price':442000},
    3000:  {'title':'ЦКТ 3000 л','diameter':1200,'height_cyl':1900,'height_cone':860,'height_full':2960,'full_volume':2473,'working_volume':1978,'wall':3,'jackets':2,'weight':350,'pressure':2.5,'from_price':598000},
    4000:  {'title':'ЦКТ 4000 л','diameter':1300,'height_cyl':2100,'height_cone':930,'height_full':3230,'full_volume':3199,'working_volume':2559,'wall':4,'jackets':2,'weight':460,'pressure':2.5,'from_price':767000},
    5000:  {'title':'ЦКТ 5000 л','diameter':1400,'height_cyl':2300,'height_cone':1000,'height_full':3500,'full_volume':4054,'working_volume':3243,'wall':4,'jackets':2,'weight':550,'pressure':2.5,'from_price':936000},
    6000:  {'title':'ЦКТ 6000 л','diameter':1500,'height_cyl':2400,'height_cone':1070,'height_full':3670,'full_volume':4871,'working_volume':3897,'wall':4,'jackets':2,'weight':650,'pressure':2.5,'from_price':1105000},
    7500:  {'title':'ЦКТ 7500 л','diameter':1550,'height_cyl':2550,'height_cone':1110,'height_full':3860,'full_volume':5800,'working_volume':4640,'wall':4,'jackets':2,'weight':730,'pressure':2.5,'from_price':1274000},
    8000:  {'title':'ЦКТ 8000 л','diameter':1600,'height_cyl':2700,'height_cone':1140,'height_full':4040,'full_volume':6193,'working_volume':4954,'wall':4,'jackets':2,'weight':810,'pressure':2.5,'from_price':1404000},
    10000: {'title':'ЦКТ 10000 л','diameter':1800,'height_cyl':2700,'height_cone':1290,'height_full':4190,'full_volume':7965,'working_volume':6372,'wall':4,'jackets':3,'weight':980,'pressure':2.5,'from_price':1755000},
    12000: {'title':'ЦКТ 12000 л','diameter':1900,'height_cyl':2800,'height_cone':1360,'height_full':4360,'full_volume':9224,'working_volume':7379,'wall':5,'jackets':3,'weight':1150,'pressure':2.5,'from_price':2054000},
    15000: {'title':'ЦКТ 15000 л','diameter':2000,'height_cyl':3200,'height_cone':1430,'height_full':4830,'full_volume':11551,'working_volume':9241,'wall':5,'jackets':3,'weight':1400,'pressure':2.5,'from_price':2496000},
    20000: {'title':'ЦКТ 20000 л','diameter':2200,'height_cyl':3600,'height_cone':1570,'height_full':5370,'full_volume':15674,'working_volume':12539,'wall':5,'jackets':4,'weight':1800,'pressure':2.5,'from_price':3224000},
    25000: {'title':'ЦКТ 25000 л','diameter':2400,'height_cyl':3700,'height_cone':1710,'height_full':5610,'full_volume':19317,'working_volume':15454,'wall':5,'jackets':4,'weight':2200,'pressure':2.5,'from_price':4030000},
    30000: {'title':'ЦКТ 30000 л','diameter':2600,'height_cyl':3900,'height_cone':1860,'height_full':5960,'full_volume':23998,'working_volume':19198,'wall':6,'jackets':4,'weight':2800,'pressure':2.5,'from_price':4875000},
    40000: {'title':'ЦКТ 40000 л','diameter':2800,'height_cyl':4400,'height_cone':2000,'height_full':6600,'full_volume':31198,'working_volume':24958,'wall':6,'jackets':4,'weight':3500,'pressure':2.5,'from_price':6240000},
    50000: {'title':'ЦКТ 50000 л','diameter':3000,'height_cyl':4800,'height_cone':2140,'height_full':7140,'full_volume':38971,'working_volume':31177,'wall':6,'jackets':4,'weight':4200,'pressure':2.5,'from_price':7670000},
    60000: {'title':'ЦКТ 60000 л','diameter':3200,'height_cyl':5000,'height_cone':2290,'height_full':7490,'full_volume':46351,'working_volume':37081,'wall':6,'jackets':4,'weight':5000,'pressure':2.5,'from_price':9360000},
    80000: {'title':'ЦКТ 80000 л','diameter':3500,'height_cyl':5400,'height_cone':2500,'height_full':8100,'full_volume':59972,'working_volume':47978,'wall':8,'jackets':4,'weight':6500,'pressure':2.5,'from_price':12350000},
    100000:{'title':'ЦКТ 100000 л','diameter':3800,'height_cyl':5500,'height_cone':2710,'height_full':8410,'full_volume':72621,'working_volume':58097,'wall':8,'jackets':4,'weight':8000,'pressure':2.5,'from_price':15600000},
    120000:{'title':'ЦКТ 120000 л','diameter':4100,'height_cyl':5700,'height_cone':2900,'height_full':8600,'full_volume':87600,'working_volume':70080,'wall':8,'jackets':4,'weight':9500,'pressure':2.5,'from_price':18720000},
    150000:{'title':'ЦКТ 150000 л','diameter':4400,'height_cyl':6000,'height_cone':3100,'height_full':9100,'full_volume':106900,'working_volume':85520,'wall':10,'jackets':4,'weight':11500,'pressure':2.5,'from_price':23400000},
    200000:{'title':'ЦКТ 200000 л','diameter':4800,'height_cyl':6500,'height_cone':3400,'height_full':9900,'full_volume':138100,'working_volume':110480,'wall':10,'jackets':4,'weight':15000,'pressure':2.5,'from_price':31200000},
}

rows_cct = []
for vol in sorted(cct_data):
    d = cct_data[vol]
    rows_cct.append([
        'Пиво', d['title'], vol,
        d['diameter'], d['height_full'], d['wall'],
        d['weight'], '', d['full_volume'], d['working_volume'],
        d['jackets'], d['pressure'], d['from_price'], ''
    ])

# ── BEER – Brew House ─────────────────────────────────────────────────
brew_data = {
    'mash-tun':       ('Заторный аппарат', {250:[600,900,3,80,9,273000,254,200],500:[750,1100,3,120,12,403000,486,400],1000:[950,1400,3,190,18,637000,992,900],2000:[1200,1700,4,310,24,1014000,1923,1700],3000:[1400,1900,4,440,30,1365000,2925,2600],5000:[1600,2300,4,640,42,1950000,4624,4200]}),
    'combined-kettle':('Заторно-сусловарочный аппарат (ЗСА)', {250:[700,800,3,90,9,325000,308,300],500:[850,1000,3,135,12,481000,567,500],1000:[1050,1300,3,210,18,741000,1126,1000],2000:[1300,1600,4,340,24,1157000,2124,1900],3000:[1500,1800,4,480,30,1560000,3181,2900],5000:[1800,2200,4,700,42,2275000,5598,5000]}),
    'lauter-tun':     ('Фильтрационный аппарат', {250:[700,750,3,75,3,247000,289,300],500:[850,950,3,115,4,377000,539,500],1000:[1100,1200,3,180,5.5,598000,1140,1000],2000:[1400,1500,4,290,7.5,936000,2309,2100],3000:[1600,1700,4,410,7.5,1274000,3418,3100],5000:[1900,2000,4,600,11,1820000,5671,5100]}),
    'brew-kettle':    ('Сусловарочный аппарат', {250:[600,950,3,70,12,234000,269,200],500:[750,1150,3,110,18,358000,508,500],1000:[950,1450,3,175,30,572000,1028,900],2000:[1200,1800,4,280,45,910000,2036,1800],3000:[1400,2000,4,400,60,1248000,3079,2800],5000:[1600,2400,4,580,90,1794000,4825,4300]}),
    'whirlpool':      ('Гидроциклонный аппарат (Вирпул)', {250:[800,700,3,55,0,169000,352,300],500:[1000,850,3,85,0,260000,668,600],1000:[1200,1100,3,135,0,416000,1244,1100],2000:[1500,1400,4,220,0,663000,2474,2200],3000:[1700,1600,4,310,0,897000,3632,3300],5000:[2000,1900,4,470,0,1300000,5969,5400]}),
    'wort-receiver':  ('Суслосборник', {500:[700,1300,3,90,0,221000,500,400],1000:[900,1500,3,145,0,351000,954,900],2000:[1100,1900,4,240,0,559000,1806,1600],3000:[1300,2100,4,330,0,754000,2787,2500],5000:[1500,2600,4,490,0,1118000,4595,4100]}),
}

rows_brew = []
for cat_key, (name, vols) in brew_data.items():
    for vol in sorted(vols):
        d = vols[vol]
        rows_brew.append([
            'Пиво', f'{name} ({cat_key})', vol,
            d[0], d[1], d[2], d[3], d[4], d[6], d[7], '', '', d[5], ''
        ])

# ── BEER – Extra ──────────────────────────────────────────────────────
beer_extra = {
    'hot-water-tank': ('Бак горячей воды (БГВ)', {500:[700,1300,3,85,12,189000,500,450],1000:[900,1500,3,135,18,286000,954,859],1500:[1000,1700,3,180,24,377000,1335,1200],2000:[1100,1900,3,230,30,468000,1806,1625],3000:[1200,2300,4,320,45,624000,2601,2341],4000:[1300,2600,4,410,60,780000,3451,3106],5000:[1400,2800,4,500,75,936000,4310,3879],6000:[1500,3000,4,590,90,1092000,5301,4771],8000:[1600,3500,4,750,110,1365000,7037,6333],10000:[1800,3500,4,900,130,1625000,8906,8015],15000:[2000,4200,5,1300,180,2210000,13195,11876],20000:[2200,4800,5,1700,240,2795000,18246,16421]}),
    'grain-mill':     ('Дробилка солода', {100:[0,0,0,35,1.5,65000],200:[0,0,0,55,2.2,95000],300:[0,0,0,75,3,130000],500:[0,0,0,110,4,190000],1000:[0,0,0,180,7.5,310000]}),
    'steam-generator':('Парогенератор', {20:[0,0,0,45,15,85000],100:[0,0,0,90,75,180000],150:[0,0,0,120,110,240000],300:[0,0,0,190,220,380000],400:[0,0,0,240,300,480000],700:[0,0,0,340,520,700000]}),
    'chiller':        ('Чиллер', {8:[0,0,0,80,3.5,160000],12:[0,0,0,110,5,220000],20:[0,0,0,160,8,340000],50:[0,0,0,300,18,580000],150:[0,0,0,600,55,1200000]}),
    'unitank':        ('Форфас (Bright Beer Tank)', {250:[500,1300,3,55,0,111000,255,200],500:[650,1600,3,85,0,169000,531,400],1000:[800,2000,3,140,0,273000,1005,900],1500:[900,2400,3,190,0,364000,1527,1300],2000:[1000,2600,4,250,0,468000,2042,1800],3000:[1200,2800,4,340,0,624000,3167,2800],5000:[1400,3200,4,500,0,936000,4926,4400]}),
    'heat-exchanger': ('Теплообменник пластинчатый', {300:[0,0,0,15,0,35000],600:[0,0,0,25,0,55000],1000:[0,0,0,35,0,85000],3000:[0,0,0,60,0,150000],5000:[0,0,0,90,0,230000]}),
}

rows_extra = []
for cat_key, (name, vols) in beer_extra.items():
    for vol in sorted(vols):
        d = vols[vol]
        is_tank = len(d) >= 9
        rows_extra.append([
            'Пиво', name, vol,
            d[0] if is_tank else '', d[1] if is_tank else '',
            d[2] if is_tank else '', d[3], d[4],
            d[6] if is_tank else '', d[7] if is_tank else '',
            '', '', d[5] if is_tank or len(d) == 7 else d[6] if len(d) > 6 else d[5], ''
        ])


# ═══════════════════════════════════════════════════════════════════
#  2. WINE
# ═══════════════════════════════════════════════════════════════════
wine_spec_ref = {
    'red-fermentation':  [500, {'diameter':800,'height':1400,'wall':2,'weight':85,'power':4,'price':215000}],
    'white-fermentation': [500, {'diameter':700,'height':1500,'wall':2,'weight':80,'power':4,'price':210000}],
    'storage-aging':      [5000,{'diameter':1500,'height':3000,'wall':3,'weight':350,'power':0,'price':475000}],
    'cold-stabilization': [1000,{'diameter':900,'height':1600,'wall':3,'weight':130,'power':6,'price':335000}],
    'blending':           [1000,{'diameter':900,'height':1500,'wall':2,'weight':105,'power':3,'price':275000}],
    'sulfitation':        [500, {'diameter':600,'height':1300,'wall':3,'weight':85,'power':0,'price':230000}],
    'universal-tank':     [1000,{'diameter':900,'height':1500,'wall':2,'weight':115,'power':5,'price':300000}],
}

wine_names = {
    'red-fermentation':  'Ферментационный танк для красных вин',
    'white-fermentation': 'Ферментационный танк для белых вин',
    'storage-aging':      'Ёмкость для выдержки и хранения вина',
    'cold-stabilization': 'Танк холодной стабилизации (криостат)',
    'blending':           'Ёмкость для купажирования вина',
    'sulfitation':        'Ёмкость сульфитации вина',
    'universal-tank':     'Винификатор (универсальный терморегулируемый танк)',
}

wine_volumes = {
    'red-fermentation':  [500,1000,1500,2000,3000,4000,5000,6000,6300,8000,10000,12500,15000,16000,20000,25000,30000,31500,40000,50000],
    'white-fermentation': [500,1000,1500,2000,3000,4000,5000,6000,6300,8000,10000,12500,15000,16000,20000,25000,31500],
    'storage-aging':      [1000,1500,2000,3000,4000,5000,6300,8000,10000,12500,15000,16000,20000,25000,30000,31500,40000,50000,63000,80000,100000,125000,160000,200000],
    'cold-stabilization': [500,1000,1500,2000,3000,4000,5000,6000,6300,8000,10000,12500,15000,16000,20000,25000,31500],
    'blending':           [1000,1500,2000,3000,4000,5000,6000,6300,8000,10000,12500,15000,16000,20000,25000,30000,31500,40000,50000],
    'sulfitation':        [500,1000,1500,2000,3000,4000,5000,6000,6300,8000,10000,12500,15000,16000,20000,25000,31500],
    'universal-tank':     [500,1000,1500,2000,3000,4000,5000,6000,6300,8000,10000,12500,15000,16000,20000,25000,30000,31500,40000,50000],
}

rows_wine = []
for cat_key in wine_spec_ref:
    ref_vol, ref_spec = wine_spec_ref[cat_key]
    name = wine_names[cat_key]
    for vol in wine_volumes[cat_key]:
        s = _specs(ref_spec, vol, ref_vol, 1.3)
        rows_wine.append([
            'Вино', name, vol,
            s['diameter'], s['height'], s['wall'],
            s['weight'], s['power'], s['full_volume'], s['working_volume'],
            '', '', s['price'], ''
        ])


# ═══════════════════════════════════════════════════════════════════
#  3. DAIRY
# ═══════════════════════════════════════════════════════════════════
dairy_spec_ref = {
    'reception':      [1000,{'diameter':900,'height':1500,'wall':2,'weight':95,'power':0,'price':150000}],
    'cooler':         [1000,{'diameter':900,'height':1500,'wall':2,'weight':130,'power':3,'price':250000}],
    'storage':        [5000,{'diameter':1400,'height':2900,'wall':3,'weight':320,'power':0,'price':420000}],
    'vdp':            [200, {'diameter':500,'height':900,'wall':2,'weight':55,'power':6,'price':130000}],
    'fermentation':   [500, {'diameter':650,'height':1400,'wall':2,'weight':75,'power':4,'price':190000}],
    'cheese-maker':   [500, {'diameter':700,'height':1200,'wall':2,'weight':90,'power':6,'price':235000}],
    'cottage-cheese': [500, {'diameter':700,'height':1200,'wall':2,'weight':95,'power':9,'price':265000}],
    'yeast':          [50,  {'diameter':300,'height':600,'wall':2,'weight':20,'power':1.5,'price':60000}],
    'brine':          [200, {'diameter':500,'height':800,'wall':3,'weight':35,'power':0,'price':75000}],
}

dairy_names = {
    'reception':      'Ёмкость приёмки молока',
    'cooler':         'Резервуар-охладитель молока',
    'storage':        'Резервуар для хранения молока',
    'vdp':            'Ванна длительной пастеризации (ВДП)',
    'fermentation':   'Ферментационный танк',
    'cheese-maker':   'Сыроизготовитель',
    'cottage-cheese': 'Творогоизготовитель',
    'yeast':          'Заквасочник',
    'brine':          'Контейнер для соления сыра',
}

dairy_volumes = {
    'reception':      [1000,1500,2000,3000,4000,5000,6000,6300,8000,10000,12500,15000,16000,20000,25000,30000,31500,40000,50000],
    'cooler':         [1000,1500,2000,3000,4000,5000,6000,6300,8000,10000,12500,15000,16000,20000,25000,30000,31500,40000,50000],
    'storage':        [3000,4000,5000,6300,8000,10000,12500,15000,16000,20000,25000,30000,31500,40000,50000,63000,80000,100000,125000,160000,200000],
    'vdp':            [200,300,500,1000,1500,2000,3000,4000,5000,6300,8000,10000],
    'fermentation':   [500,1000,1500,2000,3000,4000,5000,6000,6300,8000,10000,12500,15000,16000,20000,25000,30000,31500,40000,50000],
    'cheese-maker':   [200,300,500,1000,1500,2000,3000,4000,5000,6300,8000,10000],
    'cottage-cheese': [200,300,500,1000,1500,2000,3000,4000,5000,6300],
    'yeast':          [50,100,200,300,500,630,800,1000],
    'brine':          [200,300,500,1000,1500,2000,3000,4000,5000,6300],
}

rows_dairy = []
for cat_key in dairy_spec_ref:
    ref_vol, ref_spec = dairy_spec_ref[cat_key]
    name = dairy_names[cat_key]
    for vol in dairy_volumes[cat_key]:
        s = _specs(ref_spec, vol, ref_vol, 1.3)
        rows_dairy.append([
            'Молочка', name, vol,
            s['diameter'], s['height'], s['wall'],
            s['weight'], s['power'], s['full_volume'], s['working_volume'],
            '', '', s['price'], ''
        ])


# ═══════════════════════════════════════════════════════════════════
#  4. INDUSTRIAL
# ═══════════════════════════════════════════════════════════════════
ind_spec_ref = {
    'storage':  [5000,{'diameter':1600,'height':2800,'wall':3,'weight':320,'power':0,'price':390000}],
    'mixing':   [1000,{'diameter':1000,'height':1400,'wall':3,'weight':130,'power':7.5,'price':295000}],
    'thermal':  [2000,{'diameter':1200,'height':2000,'wall':3,'weight':200,'power':15,'price':450000}],
    'pressure': [2000,{'diameter':1100,'height':2200,'wall':5,'weight':280,'power':0,'price':490000}],
}

ind_names = {
    'storage':  'Резервуар для хранения',
    'mixing':   'Ёмкость с мешалкой',
    'thermal':  'Ёмкость с терморегуляцией',
    'pressure': 'Емкость под давлением',
}

ind_volumes = {
    'storage':  [1000,1500,2000,3000,4000,5000,6000,6300,8000,10000,12500,15000,16000,20000,25000,30000,31500,40000,50000,63000,80000,100000],
    'mixing':   [500,1000,1500,2000,3000,4000,5000,6000,6300,8000,10000,12500,15000,16000,20000,25000,30000,31500,40000,50000],
    'thermal':  [500,1000,1500,2000,3000,4000,5000,6000,6300,8000,10000,12500,15000,16000,20000,25000,30000,31500,40000,50000],
    'pressure': [200,300,400,500,1000,1500,2000,3000,4000,5000,6000,6300,8000,10000,12500,15000,16000,20000,25000,30000],
}

rows_ind = []
for cat_key in ind_spec_ref:
    ref_vol, ref_spec = ind_spec_ref[cat_key]
    name = ind_names[cat_key]
    for vol in ind_volumes[cat_key]:
        s = _specs(ref_spec, vol, ref_vol, 1.3)
        rows_ind.append([
            'Промышленка', name, vol,
            s['diameter'], s['height'], s['wall'],
            s['weight'], s['power'], s['full_volume'], s['working_volume'],
            '', '', s['price'], ''
        ])


# ═══════════════════════════════════════════════════════════════════
#  BUILD EXCEL
# ═══════════════════════════════════════════════════════════════════

SHEETS = [
    ('CCT (пиво)', rows_cct),
    ('Варочный порядок', rows_brew),
    ('Другое пивное (экстра)', rows_extra),
    ('Виноделие', rows_wine),
    ('Молочка', rows_dairy),
    ('Промышленка', rows_ind),
]

# — main sheet "Сводка" with all rows
ws_all = wb.active
write_sheet(ws_all, 'Все оборудование',
    rows_cct + rows_brew + rows_extra + rows_wine + rows_dairy + rows_ind)

# — individual category sheets
for sheet_title, sheet_rows in SHEETS:
    ws = wb.create_sheet()
    write_sheet(ws, sheet_title, sheet_rows)

path = '/Users/tretyakov/Desktop/Апгрейд/Каталог_оборудования.xlsx'
wb.save(path)
print(f'Excel saved: {path}')
print(f'Total rows: {len(rows_cct)+len(rows_brew)+len(rows_extra)+len(rows_wine)+len(rows_dairy)+len(rows_ind)}')
