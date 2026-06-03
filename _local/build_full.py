#!/usr/bin/env python3
"""
Rebuild: extract all volumes properly, compute target prices for EVERY SKU
"""
import os, re, urllib.request, ssl
ssl._create_default_https_context = ssl._create_unverified_context
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

UA = 'Mozilla/5.0'

def fetch(url):
    try:
        req = urllib.request.Request(url, headers={'User-Agent': UA})
        resp = urllib.request.urlopen(req, timeout=10)
        return resp.read().decode('utf-8')
    except: return None

def parse_price(txt):
    txt = txt.strip().replace(' ', '').replace('\u2009','').replace(',','.')
    if not txt or txt in ('-','—',''): return None
    mult = 1
    if 'млн' in txt: mult = 1_000_000; txt = txt.replace('млн','')
    elif 'тыс' in txt: mult = 1_000; txt = txt.replace('тыс','')
    txt = txt.replace('₽','').replace('руб','').strip()
    try: return float(txt) * mult
    except: return None

def fmt_full(v):
    if v is None: return '—'
    if v >= 1_000_000: return f"{v/1_000_000:.1f} млн ₽".replace('.0 ',' ')
    elif v >= 1_000: return f"{int(round(v/1000)):,} тыс ₽".replace(',',' ')
    else: return f"{int(v):,} ₽".replace(',',' ')

def extract_all_volumes(cat_url):
    """Fetch category page and extract ALL volume→price from JSON-LD"""
    html = fetch(cat_url)
    if not html: return {}
    results = {}
    # Extract JSON-LD
    jm = re.search(r'<script[^>]*type="application/ld\+json"[^>]*>(.*?)</script>', html, re.DOTALL)
    if jm:
        import json
        try:
            data = json.loads(jm.group(1))
            # Handle ItemList with ListItem elements
            items = []
            if isinstance(data, dict):
                if 'itemListElement' in data:
                    items = data['itemListElement']
                elif '@graph' in data:
                    for item in data['@graph']:
                        if 'itemListElement' in item:
                            items = item['itemListElement']
            for item in items:
                if isinstance(item, dict) and 'offers' in item:
                    name = item.get('name', '')
                    price_str = item['offers'].get('price', '0')
                    # Extract volume from name (e.g., "Ванна ... 200 л")
                    vm = re.search(r'(\d+)\s*[лlL]', name)
                    if vm:
                        vol = int(vm.group(1))
                        try:
                            pr = float(price_str)
                            if pr > 0:
                                results[vol] = pr
                        except: pass
                        else: pass
        except: pass
    
    if results:
        return results
    
    # Fallback: try to find data in JSON-like attributes
    for m in re.finditer(r'data-price="([^"]*)"[^>]*data-name="([^"]*)"', html):
        name = m.group(2)
        vm = re.search(r'(\d+)\s*[лlL]', name)
        if vm:
            vol = int(vm.group(1))
            try: pr = float(m.group(1))
            except: continue
            if pr > 0: results[vol] = pr
    
    # Fallback 2: try to find structured data around volume links
    # Look for volume link followed by a price element
    for m in re.finditer(
        r'href="([^"]*/(\d+)[lL]/)"[^>]*>.*?от\s*([\d\s]+)\s*(тыс|млн|₽)',
        html, re.DOTALL
    ):
        vol = int(m.group(2))
        pr = parse_price(m.group(3).strip() + ' ' + m.group(4).strip())
        if pr: results[vol] = pr
    
    return results

# ═══ ALL CATEGORY PAGES TO CRAWL ═══
CATALOG = [
    ('Пиво', 'ЦКТ', '/catalog/beer/cct/'),
    ('Пиво', 'Заторные чаны', '/catalog/beer/brew-house/mash-tun/'),
    ('Пиво', 'Фильтрационные чаны', '/catalog/beer/brew-house/lauter-tun/'),
    ('Пиво', 'Сусловарочные котлы', '/catalog/beer/brew-house/brew-kettle/'),
    ('Пиво', 'Вильпули', '/catalog/beer/brew-house/whirlpool/'),
    ('Пиво', 'Сборники сусла', '/catalog/beer/brew-house/wort-receiver/'),
    ('Пиво', 'Баки горячей воды', '/catalog/beer/hot-water-tank/'),
    ('Пиво', 'Форфасы', '/catalog/beer/unitank/'),
    ('Молоко', 'ВДП', '/catalog/dairy/vdp/'),
    ('Молоко', 'Охладители молока', '/catalog/dairy/cooler/'),
    ('Молоко', 'Сыроизготовители', '/catalog/dairy/cheese-maker/'),
    ('Молоко', 'Резервуары хранения', '/catalog/dairy/storage/'),
    ('Молоко', 'Ферментационные танки', '/catalog/dairy/fermentation/'),
    ('Молоко', 'Творогоизготовители', '/catalog/dairy/cottage-cheese/'),
    ('Молоко', 'Заквасочники', '/catalog/dairy/yeast/'),
    ('Молоко', 'Контейнеры для соления', '/catalog/dairy/brine/'),
    ('Молоко', 'Ёмкости приёмки', '/catalog/dairy/reception/'),
    ('Вино', 'Красная ферментация', '/catalog/wine/red-fermentation/'),
    ('Вино', 'Белая ферментация', '/catalog/wine/white-fermentation/'),
    ('Вино', 'Выдержка и хранение', '/catalog/wine/storage-aging/'),
    ('Вино', 'Холодная стабилизация', '/catalog/wine/cold-stabilization/'),
    ('Вино', 'Купажирование', '/catalog/wine/blending/'),
    ('Вино', 'Сульфитация', '/catalog/wine/sulfitation/'),
    ('Вино', 'Винификатор', '/catalog/wine/universal-tank/'),
    ('Пром', 'Резервуары хранения', '/catalog/industrial/storage/'),
    ('Пром', 'Ёмкости с мешалкой', '/catalog/industrial/mixing/'),
    ('Пром', 'Ёмкости с терморегуляцией', '/catalog/industrial/thermal/'),
    ('Пром', 'Ёмкости под давлением', '/catalog/industrial/pressure/'),
]

print("Fetching all category pages...")
all_data = {}
for section, cat, url_path in CATALOG:
    url = 'https://ob-kub.ru' + url_path
    volumes = extract_all_volumes(url)
    all_data[(section, cat)] = volumes
    print(f"  {section} > {cat}: {len(volumes)} volumes")

# ═══ COMPETITOR PRICES ═══
COMP = [
    ('ЦКТ', 100, 'ТехТанк', 107000, 'est'),
    ('ЦКТ', 250, 'ТехТанк', 170000, 'est'),
    ('ЦКТ', 500, 'ТехТанк', 294000, 'open'),
    ('ЦКТ', 1000, 'ТехТанк', 362335, 'open'),
    ('ЦКТ', 1000, 'АгроДеталь', 439000, 'open'),
    ('ЦКТ', 1000, 'Доля Ангелов', 550000, 'open'),
    ('ЦКТ', 1500, 'ТехТанк', 420000, 'est'),
    ('ЦКТ', 2000, 'ТехТанк', 478600, 'open'),
    ('ЦКТ', 2000, 'АгроДеталь', 574000, 'open'),
    ('ЦКТ', 3000, 'ТехТанк', 560650, 'open'),
    ('ЦКТ', 3000, 'АгроДеталь', 720000, 'open'),
    ('ЦКТ', 4000, 'ТехТанк', 650000, 'est'),
    ('ЦКТ', 5000, 'ТехТанк', 773470, 'open'),
    ('ЦКТ', 6000, 'ТехТанк', 900000, 'est'),
    ('ЦКТ', 7500, 'ТехТанк', 1000000, 'est'),
    ('ЦКТ', 8000, 'ТехТанк', 1100000, 'est'),
    ('ЦКТ', 10000, 'ТехТанк', 1045600, 'open'),
    ('ЦКТ', 10000, 'АгроДеталь', 1720000, 'open'),
    ('ЦКТ', 12000, 'ТехТанк', 1200000, 'est'),
    ('ЦКТ', 15000, 'ТехТанк', 1400000, 'est'),
    ('ЦКТ', 20000, 'ТехТанк', 1800000, 'est'),
    ('ЦКТ', 25000, 'ТехТанк', 2200000, 'est'),
    ('ЦКТ', 30000, 'ТехТанк', 2800000, 'est'),
    ('ЦКТ', 40000, 'ТехТанк', 3600000, 'est'),
    ('ЦКТ', 50000, 'ТехТанк', 4500000, 'est'),
    ('ЦКТ', 60000, 'ТехТанк', 5500000, 'est'),
    ('ЦКТ', 80000, 'ТехТанк', 7200000, 'est'),
    ('ЦКТ', 100000, 'ТехТанк', 9000000, 'est'),
    ('ЦКТ', 120000, 'ТехТанк', 11000000, 'est'),
    ('ЦКТ', 150000, 'ТехТанк', 14000000, 'est'),
    ('ЦКТ', 200000, 'ТехТанк', 18500000, 'est'),
    ('ВДП', 200, 'Агрокомплект', 180000, 'est'),
    ('ВДП', 500, 'Агрокомплект', 280000, 'est'),
    ('ВДП', 1000, 'Агрокомплект', 368900, 'open'),
    ('ВДП', 1500, 'Агрокомплект', 480000, 'est'),
    ('ВДП', 2000, 'АгроДеталь', 696000, 'open'),
    ('ВДП', 3000, 'АгроДеталь', 833000, 'open'),
    ('ВДП', 4000, 'АгроДеталь', 900000, 'est'),
    ('ВДП', 5000, 'АгроДеталь', 1100000, 'est'),
    ('ВДП', 6300, 'АгроДеталь', 1300000, 'est'),
    ('ВДП', 8000, 'АгроДеталь', 1500000, 'est'),
    ('ВДП', 10000, 'АгроДеталь', 1700000, 'est'),
    ('Сыроизг', 200, 'ТехТанк', 385952, 'open'),
    ('Сыроизг', 500, 'ТехТанк', 507455, 'open'),
    ('Сыроизг', 1000, 'АгроДеталь', 825000, 'open'),
    ('Сыроизг', 10000, 'АгроДеталь', 2348000, 'open'),
    ('Охл', 1000, 'ТехТанк', 549940, 'open'),
    ('Охл', 5000, 'ТехТанк', 1098890, 'open'),
    ('Охл', 10000, 'ТехТанк', 1897870, 'open'),
    ('Охл', 10000, 'АгроДеталь', 1750000, 'open'),
]

# Build best-prices lookup per (category, volume)
best = {}
for cat, vol, name, price, src in COMP:
    key = (cat, vol)
    if key not in best or price < best[key][0]:
        best[key] = (price, name, src)

def get_comp_ref(cat_comp, vol):
    """Get best reference competitor price for a volume"""
    key = (cat_comp, vol)
    if key in best: return best[key]
    known = [(v, p, n, s) for (c, v), (p, n, s) in best.items() if c == cat_comp]
    if not known: return None
    known.sort()
    if vol <= known[0][0]:
        v0, p, n, s = known[0]
        return (p * vol / v0, f"~{n}", 'est')
    if vol >= known[-1][0]:
        vN, p, n, s = known[-1]
        return (p * vol / vN, f"~{n}", 'est')
    for i in range(len(known)-1):
        if known[i][0] <= vol <= known[i+1][0]:
            v1, p1, n1, s1 = known[i]
            v2, p2, n2, s2 = known[i+1]
            r = (vol - v1) / (v2 - v1) if v2 != v1 else 0
            ip = p1 + (p2 - p1) * r
            return (ip, f"~{n1}/{n2}", 'est')
    return None

# Complexity factor vs ЦКТ for each category
CPX = {
    'ЦКТ': 1.0, 'Заторные чаны': 0.85, 'Фильтрационные чаны': 0.8,
    'Сусловарочные котлы': 0.9, 'Вильпули': 0.7, 'Сборники сусла': 0.55,
    'Баки горячей воды': 0.45, 'Форфасы': 0.85,
    'ВДП': 1.0, 'Охладители молока': 1.0, 'Сыроизготовители': 1.0,
    'Резервуары хранения': 0.6, 'Ферментационные танки': 0.85,
    'Творогоизготовители': 0.7, 'Заквасочники': 0.5,
    'Контейнеры для соления': 0.45, 'Ёмкости приёмки': 0.55,
    'Красная ферментация': 0.85, 'Белая ферментация': 0.85,
    'Выдержка и хранение': 0.6, 'Холодная стабилизация': 0.85,
    'Купажирование': 0.7, 'Сульфитация': 0.6, 'Винификатор': 1.05,
    'Ёмкости с мешалкой': 0.9, 'Ёмкости с терморегуляцией': 0.85,
    'Ёмкости под давлением': 1.15,
}

# Sheet styles
wb = Workbook()
hfont = Font(name='Arial', bold=True, size=9, color='FFFFFF')
hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
cbfont = Font(name='Arial', bold=True, size=10, color='2F5496')
cbfill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
gfill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
yfill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
rfill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
nfill = PatternFill(start_color='F8E8C8', end_color='F8E8C8', fill_type='solid')
nfont = Font(name='Arial', size=9); bfont = Font(name='Arial', bold=True, size=9)
thin = Side(style='thin'); bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical='top')
cc = Alignment(wrap_text=True, vertical='center', horizontal='center')

def hdr(ws, r, n):
    for c in range(1, n+1):
        cl = ws.cell(row=r, column=c)
        cl.font = hfont; cl.fill = hfill; cl.alignment = cc; cl.border = bdr

def sec(ws, r, t, n):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n)
    cl = ws.cell(row=r, column=1, value=t)
    cl.font = cbfont; cl.fill = cbfill; cl.alignment = Alignment(vertical='center')
    for c in range(1, n+1):
        ws.cell(row=r, column=c).border = bdr; ws.cell(row=r, column=c).fill = cbfill

def add(ws, r, data, fill=None):
    for i, v in enumerate(data, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = bfont if i in (3,4,6,8) else nfont
        c.border = bdr; c.alignment = wrap
        if fill: c.fill = fill

# Direct competitor mapping (categories that map to a COMP category)
DIRECT = {'ЦКТ':'ЦКТ','Заторные чаны':'Заторный','ВДП':'ВДП','Охладители молока':'Охл','Сыроизготовители':'Сыроизг'}

# ═══════ SHEET 1: ALL ═══════
ws = wb.active; ws.title = "Все товары"
H1 = ["Раздел","Категория","Объём","Цена ob-kub","Конкурент","Цена конк.","Тип","Цель (-5%)","Измен. ₽","Измен. %","Основание"]
n1 = len(H1)
for i, hh in enumerate(H1, 1): ws.cell(row=1, column=i, value=hh)
hdr(ws, 1, n1)

r = 2
stats_raise = stats_lower = stats_calc = stats_ok = 0

for (section, cat) in sorted(all_data.keys(), key=lambda x: (x[0], x[1])):
    prices = all_data[(section, cat)]
    sec(ws, r, f"{section} > {cat}", n1); r += 1
    direct_cat = DIRECT.get(cat)
    
    for vol in sorted(prices.keys()):
        cur = prices[vol]
        # Try direct competitor first
        cinfo = get_comp_ref(direct_cat, vol) if direct_cat else None
        # If none, use ЦКТ × complexity
        if cinfo is None:
            cpx = CPX.get(cat, 1.0)
            ckt = get_comp_ref('ЦКТ', vol)
            if ckt:
                ref_p, ref_n, ref_t = ckt
                cinfo = (ref_p * cpx, f"~ЦКТ×{cpx} ({ref_n})", 'calc')
        # If still none, use ЦКТ directly
        if cinfo is None:
            cinfo = get_comp_ref('ЦКТ', vol)
            if cinfo:
                cinfo = (cinfo[0], cinfo[1], 'est')
        
        if cinfo:
            cp, cn, ct = cinfo
            target = int(cp * 0.95)
            delta = target - cur
            pct = (delta / cur) * 100
            delta_txt = f"{delta:+,}".replace(',',' ') if abs(delta) > 500 else '—'
            pct_txt = f"{pct:+.1f}%" if abs(pct) > 0.3 else '≈0%'
        else:
            target = cur; cn = '—'; ct = '—'
            delta_txt = '—'; pct_txt = '—'
        
        if ct == '—': clr = nfill; stats_ok += 1
        elif ct == 'calc': clr = nfill; stats_calc += 1
        elif pct > 5: clr = yfill; stats_raise += 1
        elif pct < -5: clr = rfill; stats_lower += 1
        else: clr = gfill; stats_ok += 1
        
        ct_txt = {'open':'ОТКР','found':'НАЙД','est':'ОЦЕН','calc':'РАСЧЁТ','—':'—'}.get(ct, ct)
        
        add(ws, r, [
            section, cat, f"{vol} л", fmt_full(cur),
            cn, fmt_full(cp) if cinfo else '—', ct_txt,
            fmt_full(target), delta_txt, pct_txt,
            {'open':'Цена конкурента','found':'Раскопана','est':'Интерполяция','calc':'ЦКТ×коэфф','—':'Нет данных'}.get(ct,'')
        ], clr)
        r += 1

for i, w in enumerate([12, 24, 12, 14, 24, 14, 8, 14, 14, 10, 36], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ═══════ SHEET 2: CHANGES ONLY ═══════
ws2 = wb.create_sheet("Что менять")
H2 = ["Раздел","Категория","Объём","Цена сейчас","Ориентир","Цена ориент.","Тип","Цель -5%","Дельта ₽","Дельта %","Действие"]
n2 = len(H2)
for i, hh in enumerate(H2, 1): ws2.cell(row=1, column=i, value=hh)
hdr(ws2, 1, n2)

r2 = 2
for (section, cat) in sorted(all_data.keys(), key=lambda x: (x[0], x[1])):
    prices = all_data[(section, cat)]
    sec(ws2, r2, f"{section} > {cat}", n2); r2 += 1
    direct_cat = DIRECT.get(cat)
    
    for vol in sorted(prices.keys()):
        cur = prices[vol]
        cinfo = get_comp_ref(direct_cat, vol) if direct_cat else None
        if cinfo is None:
            cpx = CPX.get(cat, 1.0)
            ckt = get_comp_ref('ЦКТ', vol)
            if ckt:
                cinfo = (ckt[0] * cpx, f"~ЦКТ×{cpx} ({ckt[1]})", 'calc')
        if cinfo is None:
            cinfo = get_comp_ref('ЦКТ', vol)
            if cinfo:
                cinfo = (cinfo[0], cinfo[1], 'est')
        
        if cinfo:
            cp, cn, ct = cinfo
            target = int(cp * 0.95)
            delta = target - cur
            pct = (delta / cur) * 100
            delta_txt = f"{delta:+,}".replace(',',' ') if abs(delta) > 500 else '—'
            pct_txt = f"{pct:+.1f}%" if abs(pct) > 0.3 else '≈0%'
            
            if ct == 'calc': action = '📐 РАСЧЁТ'; clr = nfill
            elif pct > 5: action = '⬆ ПОДНЯТЬ'; clr = yfill
            elif pct < -5: action = '⬇ СНИЗИТЬ'; clr = rfill
            else: action = '🟢 OK'; clr = gfill
            
            ct_txt = {'open':'ОТКР','found':'НАЙД','est':'ОЦЕН','calc':'РАСЧЁТ'}.get(ct, ct)
            
            ws2.cell(row=r2, column=1, value=section).font = nfont
            ws2.cell(row=r2, column=2, value=cat).font = nfont
            ws2.cell(row=r2, column=3, value=f"{vol} л").font = nfont
            ws2.cell(row=r2, column=4, value=fmt_full(cur)).font = nfont
            ws2.cell(row=r2, column=5, value=cn).font = nfont
            ws2.cell(row=r2, column=6, value=fmt_full(cp)).font = bfont
            ws2.cell(row=r2, column=7, value=ct_txt).font = nfont
            ws2.cell(row=r2, column=8, value=fmt_full(target)).font = bfont
            ws2.cell(row=r2, column=9, value=delta_txt).font = bfont
            ws2.cell(row=r2, column=10, value=pct_txt).font = bfont
            ws2.cell(row=r2, column=11, value=action).font = bfont
            for c in range(1, n2+1):
                ws2.cell(row=r2, column=c).border = bdr
                ws2.cell(row=r2, column=c).fill = clr
            r2 += 1

for i, w in enumerate([12, 24, 12, 14, 24, 14, 8, 14, 14, 12, 14], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'

path = "/Users/tretyakov/Desktop/Апгрейд/Все_цены_и_предложения_ob-kub.xlsx"
wb.save(path)
total = sum(len(v) for v in all_data.values())
print(f"\nOK: {path}")
print(f"Sheet 1: {r-1} rows ({total} товаров)")
print(f"Sheet 2: {r2-1} rows")
print(f"\nСтатистика:")
print(f"  ⬆ ПОДНЯТЬ: {stats_raise}")
print(f"  ⬇ СНИЗИТЬ: {stats_lower}")
print(f"  📐 РАСЧЁТ: {stats_calc}")
print(f"  🟢 OK: {stats_ok}")
print(f"  Всего: {stats_raise + stats_lower + stats_calc + stats_ok}")
